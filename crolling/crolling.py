import re, openpyxl
from openpyxl import Workbook
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
from gensim.summarization.summarizer import summarize
from newspaper import Article
import requests
from PIL import Image
import urllib.request
import time
from io import BytesIO
import kss

wb = Workbook()
ws_naver_news = wb.create_sheet()
ws_naver_news.title = 'naver news'

ws_wiki_news = wb.create_sheet("wiki news")


driver = webdriver.Chrome(r"C:\Users\PC\chromedriver.exe")

search = input('검색어를 입력하세요 : ')
url = 'https://search.naver.com/search.naver?where=news&sm=tab_jum&query='+search

original_html = requests.get(url)
html = BeautifulSoup(original_html.text, "html.parser")

articles = html.select("div.group_news > ul.list_news > li div.news_area > a")



ws_naver_news.append(["기사 타이틀", "관련 URL"])

ws_naver_news.column_dimensions['A'].width = 100
ws_naver_news.column_dimensions['B'].width = 12


A1_cell = ws_naver_news['A1']
A1_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
A1_cell.font = openpyxl.styles.Font(color='0055FF')

B1_cell = ws_naver_news['B1']
B1_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
B1_cell.font = openpyxl.styles.Font(color='0055FF')

print(url)

news_title = []
news_url = []

for i in articles:
    news_title.append(i.attrs['title'])
    
for j in articles:
    news_url.append(j.attrs['href'])
    
              

idx = 2
for title,url in zip(news_title, news_url):
    p_title = title
    p_link = url
    
    ws_naver_news.append([p_title, "이동하기"])
    ws_naver_news.cell(row=idx, column=2).hyperlink = p_link
    
    idx += 1
    

      
wb.save('sample_wow.xlsx')
wb.close
